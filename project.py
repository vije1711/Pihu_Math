# Importing necessary libraries and modules
import random
import re
from collections import Counter
import os
import sys
import json
import pyttsx3
from tkinter import *
from tkinter import messagebox, filedialog
from datetime import datetime
from fpdf import FPDF
import pandas as pd
import numpy as np
import math
from openpyxl import Workbook, load_workbook
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# Determine the directory where the script or executable is running and
# where user-selected paths should be stored
if getattr(sys, 'frozen', False):
    resource_dir = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    script_dir = os.path.dirname(sys.executable)
else:
    resource_dir = os.path.dirname(os.path.abspath(__file__))
    script_dir = resource_dir

SAVE_FILE = os.path.join(script_dir, ".save_path.txt")


def resource_path(filename: str) -> str:
    """Return absolute path to a resource file."""
    return os.path.join(resource_dir, filename)


def get_output_dir():
    """Return a writable directory for output files."""
    path = None
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r") as fh:
            p = fh.read().strip()
        if p and os.path.isdir(p) and os.access(p, os.W_OK):
            path = p
    if not path:
        root = Tk()
        root.withdraw()
        chosen = ""
        while not chosen or not os.path.isdir(chosen):
            chosen = filedialog.askdirectory(title="Select folder to save files")
            if not chosen:
                continue
        root.destroy()
        os.makedirs(chosen, exist_ok=True)
        with open(SAVE_FILE, "w") as fh:
            fh.write(chosen)
        path = chosen
    return path


# Folder where all generated files will be saved
OUTPUT_DIR = get_output_dir()

# --- Adaptive difficulty settings ---
DIFFICULTY_FILE = os.path.join(OUTPUT_DIR, "difficulty_scores.json")
DEFAULT_DIFFICULTY = {
    "+": 2.0,
    "-": 2.0,
    # start multiplication slightly easier than other operations
    "*": 1.5,
    "/": 2.0,
    "fraction": 2.0,
    "factors_primes": 2.0,
    "prime_factorization": 2.0,
    "hcf": 2.0,
    "lcm": 2.0,
}


def load_difficulty_scores():
    try:
        with open(DIFFICULTY_FILE, "r") as fh:
            return json.load(fh)
    except Exception:
        return DEFAULT_DIFFICULTY.copy()


def save_difficulty_scores(scores):
    with open(DIFFICULTY_FILE, "w") as fh:
        json.dump(scores, fh)


def write_difficulty_sheet(wb, scores):
    """Append a timestamped row of difficulty scores."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if "Difficulty" not in wb.sheetnames:
        diff_ws = wb.create_sheet("Difficulty")
        headers = ["Timestamp"] + [op_names.get(k, k) for k in scores.keys()]
        diff_ws.append(headers)
    else:
        diff_ws = wb["Difficulty"]
        if diff_ws.max_row == 0:
            headers = ["Timestamp"] + [op_names.get(k, k) for k in scores.keys()]
            diff_ws.append(headers)
    row = [timestamp] + [round(scores[k], 2) for k in scores.keys()]
    diff_ws.append(row)


def append_difficulty_session(scores):
    """Persist difficulty scores to the AllSessions workbook."""
    path = os.path.join(OUTPUT_DIR, "AllSessions.xlsx")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
    write_difficulty_sheet(wb, scores)
    wb.save(path)


difficulty_scores = load_difficulty_scores()


op_names = {
    "+": "Addition",
    "-": "Subtraction",
    "*": "Multiplication",
    "/": "Division",
    "fraction": "Fractions",
    "factors_primes": "Factors & Primes",
    "prime_factorization": "Prime Factorization",
    "hcf": "HCF",
    "lcm": "LCM",
}


def load_difficulty_history():
    path = os.path.join(OUTPUT_DIR, "AllSessions.xlsx")
    if not os.path.exists(path):
        return {}
    try:
        df = pd.read_excel(path, sheet_name="Difficulty")
    except Exception:
        return {}
    history = {}
    if set(["Operation", "Difficulty Score"]).issubset(df.columns):
        grouped = df.groupby("Operation")
        for key in difficulty_scores.keys():
            name = op_names.get(key, key)
            if name in grouped.groups:
                vals = grouped.get_group(name)["Difficulty Score"].dropna().tolist()
                history[key] = vals
    else:
        for op in difficulty_scores.keys():
            col = op_names.get(op, op)
            if col in df.columns:
                history[op] = df[col].dropna().tolist()
    return history


def compute_thresholds(history):
    stats = {}
    for op, vals in history.items():
        if vals:
            mean = float(np.mean(vals))
            std = float(np.std(vals))
        else:
            mean = difficulty_scores.get(op, 2.0)
            std = 0.0
        stats[op] = (mean, std)
    for op in difficulty_scores.keys():
        if op not in stats:
            stats[op] = (difficulty_scores.get(op, 2.0), 0.0)
    return stats


def determine_difficulty_levels(scores, stats):
    levels = {}
    for op, val in scores.items():
        mean, std = stats.get(op, (val, 0.0))
        if val < mean - std:
            levels[op] = "Easy"
        elif val > mean + std:
            levels[op] = "Hard"
        else:
            levels[op] = "Medium"
    return levels


class Exam:
    """
    Class representing a math quiz.

    Attributes:
        question (str): The math question.
        _X (int): Operand X.
        _Y (int): Operand Y.
        _Z (int): Operand Z.
        _S (str): The type of operation.
        answer_actual (int): The actual answer.
        answer_actual_remainder (int): The remainder for division problems.
        answer_user (int | list[int]): The user's answer. May be a list for
            prime factorization questions.
        answer_user_remainder (int): The user's remainder for division problems.
        _score (int): The user's score.
    """
    @classmethod
    def quiz(cls, operation, level):
        """Generate a random math question based on an operation and difficulty level."""
        S = operation
        score = difficulty_scores.get(S, 2.0)
        adjust = {"Easy": -0.5, "Medium": 0.0, "Hard": 0.5}
        difficulty = score + adjust.get(level, 0.0)
        choices = None
        limit = int(10 ** max(difficulty, 1))
        while True:
            if S == "-":
                if difficulty < 1.5:
                    X = random.randint(5, 20)
                    Y = random.randint(1, 4)
                elif difficulty < 2.5:
                    X = random.randint(30, 99)
                    Y = random.randint(10, 29)
                else:
                    X = random.randint(100, 299)
                    Y = random.randint(50, 150)
                if X > Y:
                    quiz = f"{X} - {Y}"
                    break
                continue
            elif S == "+":
                if difficulty < 1.5:
                    X = random.randint(1, 20)
                    Y = random.randint(1, 20)
                elif difficulty < 2.5:
                    X = random.randint(10, 99)
                    Y = random.randint(10, 99)
                else:
                    X = random.randint(100, 299)
                    Y = random.randint(50, 150)
                quiz = f"{X} + {Y}"
                break
            elif S == "*":
                if difficulty < 1.5:
                    X = random.randint(2, 9)
                    Y = random.randint(2, 9)
                elif difficulty < 2.5:
                    X = random.randint(5, 12)
                    Y = random.randint(3, 12)
                else:
                    X = random.randint(10, 20)
                    Y = random.randint(5, 20)
                quiz = f"{X} * {Y}"
                break
            elif S == "/":
                if difficulty < 1.5:
                    Y = random.randint(2, 5)
                    X = random.randint(Y + 1, 20)
                elif difficulty < 2.5:
                    Y = random.randint(2, 9)
                    X = random.randint(Y + 1, 99)
                else:
                    Y = random.randint(3, 12)
                    X = random.randint(Y + 1, 299)
                if X % Y != 0:
                    quiz = f"{X} / {Y}"
                    break
                continue
            elif S == "fraction":
                if difficulty < 1.5:
                    # add fractions with like denominators
                    denom = random.randint(2, 6)
                    a = random.randint(1, denom - 2)
                    b = random.randint(1, denom - a - 1)
                    X, Y, Z = a, b, denom
                    quiz = f"{a}/{denom} + {b}/{denom}"
                    answer = a + b
                    Z = (answer, denom)
                elif difficulty < 2.5:
                    # add fractions with unlike denominators
                    d1 = random.randint(2, 8)
                    d2 = random.choice([n for n in range(2, 9) if n != d1])
                    n1 = random.randint(1, d1 - 1)
                    n2 = random.randint(1, d2 - 1)
                    l = lcm_of_numbers([d1, d2])
                    total = n1 * (l // d1) + n2 * (l // d2)
                    g = math.gcd(total, l)
                    X, Y, Z = (n1, d1), (n2, d2), None
                    quiz = f"{n1}/{d1} + {n2}/{d2}"
                    Z = (total // g, l // g)
                else:
                    # simplify a fraction
                    den = random.randint(4, 20)
                    num = random.randint(2, den - 1)
                    mult = random.randint(2, 5)
                    X = num * mult
                    Y = den * mult
                    g = math.gcd(X, Y)
                    quiz = f"Simplify {X}/{Y}"
                    Z = (X // g, Y // g)
                choices = None
                break
            elif S == "factors_primes":
                if difficulty < 1.5:
                    X = random.randint(2, 20)
                elif difficulty < 2.5:
                    X = random.randint(20, 100)
                else:
                    X = random.randint(100, 300)
                quiz = f"How many factors does {X} have?"
                Z = factors_of(X)
                break
            elif S == "prime_factorization":
                while True:
                    if difficulty < 1.5:
                        X = random.randint(20, 50)
                    elif difficulty < 2.5:
                        X = random.randint(50, 150)
                    else:
                        X = random.randint(150, 300)
                    pf = prime_factorization(X)
                    if len(set(pf)) >= 2:
                        method = random.choice(["factor tree", "division"])
                        quiz = f"What are the prime factors of {X} using the {method} method?"
                        Z = pf
                        break
                break
            elif S == "hcf":
                from math import gcd
                if difficulty < 1.5:
                    count = 2
                    rng = range(2, 21)
                elif difficulty < 2.5:
                    count = 2
                    rng = range(10, 100)
                else:
                    count = 3
                    rng = range(20, 200)
                while True:
                    nums = random.sample(rng, count)
                    g = gcd(nums[0], nums[1])
                    if count == 3:
                        g = gcd(g, nums[2])
                    if g > 1:
                        break
                method = random.choice([
                    "listing factors",
                    "prime factorization",
                    "division method",
                ])
                if len(nums) == 2:
                    X, Y = nums
                    Z = None
                else:
                    X, Y, Z = nums
                if len(nums) == 3:
                    num_text = f"{nums[0]}, {nums[1]}, and {nums[2]}"
                else:
                    num_text = f"{nums[0]} and {nums[1]}"
                method_text = (
                    "by listing factors"
                    if method == "listing factors"
                    else (
                        "using prime factorization"
                        if method == "prime factorization"
                        else "using the division method"
                    )
                )
                quiz = f"Find the HCF of {num_text} {method_text}."
                obj = cls(quiz, X, Y, Z, S, choices)
                obj.numbers = nums
                obj.method = method
                return obj
            elif S == "lcm":
                if difficulty < 1.5:
                    rng = range(2, 21)
                    count = 2
                elif difficulty < 2.5:
                    rng = range(6, 41)
                    count = 2
                else:
                    rng = range(10, 60)
                    count = 3
                nums = random.sample(rng, count)
                method = random.choice([
                    "listing multiples",
                    "prime factorization",
                    "division method",
                ])
                if len(nums) == 2:
                    X, Y = nums
                    Z = None
                else:
                    X, Y, Z = nums
                if len(nums) == 3:
                    num_text = f"{nums[0]}, {nums[1]}, and {nums[2]}"
                else:
                    num_text = f"{nums[0]} and {nums[1]}"
                method_text = (
                    "by listing multiples"
                    if method == "listing multiples"
                    else (
                        "using prime factorization"
                        if method == "prime factorization"
                        else "using the division method"
                    )
                )
                quiz = f"Find the LCM of {num_text} {method_text}."
                obj = cls(quiz, X, Y, Z, S, choices)
                obj.numbers = nums
                obj.method = method
                return obj
        return cls(quiz, X, Y, Z, S, choices)

     # Initialize Exam object
    def __init__(self, question, X, Y, Z, S, choices=None):
        """
        Initialize Exam object.

        Args:
            question (str): The math question.
            X (int): Operand X.
            Y (int): Operand Y.
            Z (int | list): Operand Z or a list of factors for 'factors_primes'.
            S (str): The type of operation.
        """
        self.question = question
        self._X, self._Y, self._Z, self._S = X, Y, Z, S
        self._choices = choices
        
        # Calculate the actual answer based on the operation
        if self._S == "+":
            self.answer_actual = self._X + self._Y
        elif self._S == "-":
            self.answer_actual = self._X - self._Y
        elif self._S == "*":
            self.answer_actual = self._X * self._Y
        elif self._S == "/":
            self.answer_actual = int(self._X / self._Y)         # Applied 'int' Method for proper feedback dispaly on Answer Submission!
            self.answer_actual_remainder = self._X % self._Y
        elif self._S == "fraction":
            self.answer_actual = self._Z
        elif self._S == "factors_primes":
            self.factors = self._Z
            self.is_prime = len(self._Z) == 2
            self.twin_pair = twin_prime_pair(self._X)
            self.answer_actual = len(self._Z)
        elif self._S == "prime_factorization":
            self.answer_actual = self._Z
        elif self._S == "hcf":
            from math import gcd
            if self._Z:
                self.answer_actual = gcd(gcd(self._X, self._Y), self._Z)
            else:
                self.answer_actual = gcd(self._X, self._Y)
        elif self._S == "lcm":
            nums = [self._X, self._Y] if self._Z is None else [self._X, self._Y, self._Z]
            self.answer_actual = lcm_of_numbers(nums)
        self.answer_user = 0
        self.answer_user_remainder = 0

    @property
    def choices(self):
        return self._choices
        
    @property
    def question(self):
        """Getter method for the question."""
        return self._question

    @question.setter
    def question(self, question):
        self._question = question

    @property
    def answer_actual(self):
        return self._answer_actual
    
    @answer_actual.setter
    def answer_actual(self, answer):
        self._answer_actual = answer

    @property
    def answer_actual_remainder(self):
        return self._answer_actual_remainder
    
    @answer_actual_remainder.setter
    def answer_actual_remainder(self, answer):
        self._answer_actual_remainder = answer

    @property
    def answer_user(self):
        return self._answer_user
    
    @answer_user.setter
    def answer_user(self, answer_user):
        """Set the user's answer."""
        if isinstance(answer_user, list):
            self._answer_user = answer_user
        elif answer_user >= 0:
            self._answer_user = answer_user

    @property
    def answer_user_remainder(self):
        return self._answer_user_remainder
    
    @answer_user_remainder.setter
    def answer_user_remainder(self, answer_user):
        if answer_user >=0:
                self._answer_user_remainder = answer_user

    @property
    def score(self):
        return self._score
    
    @score.setter
    def score(self, marks):
        self._score = marks


class GUI_Exam(Exam):
    """
    Class representing the graphical user interface for a math exam application.

    Attributes:
        - root: Tkinter root window
        - exam_frame: Tkinter frame for the exam interface
        - result_frame: Tkinter frame for the result interface
        - input_user_answer: Tkinter Entry widget for user input
        - input_user_answer_remainder: Tkinter Entry widget for remainder input
        - display_question: Tkinter Label for displaying math questions
        - sound_variable: Tkinter StringVar for sound option
        - grade: Tkinter StringVar for displaying the user's grade
        - grade_label: Tkinter Label for displaying the grade
        - stat_frame: Tkinter Frame for displaying exam statistics
        - quit_button: Tkinter Button for quitting the application

    Methods:
        - __init__: Initialize the GUI_Exam object and set up the initial state.
        - launch_main: Static method to launch the main GUI window.
        - generate_question: Generate and display a new math question.
        - check_answer: Check the user's answer and provide feedback.
        - launch_result_frame: Switch to the result interface after completing the exam.
        - get_grade: Calculate the user's grade based on the exam score.
        - for_correct_answer: Return a random message for correct answers.
        - for_incorrect_answer: Return a random message for incorrect answers.
        - for_failed_attempt: Return a random message for failed attempts.
        - tell_grade: Provide a random congratulatory message based on the grade.
        - store_data: Store user data in a text file during the exam.
        - make_pdf: Generate a PDF report based on user performance.

    Note: This class inherits from Tkinter's Frame class.
    """    
    
    engine = pyttsx3.init()
    voices = engine.getProperty('voices')                               # getting details of available voices
    preferred = next((v for v in voices if "english" in v.name.lower()), voices[0])
    engine.setProperty('voice', preferred.id)                           # pick an English voice if available
    engine.setProperty('rate', 170)                                     # slightly slower for natural speech
    engine.setProperty('volume', 1.0)
    try:
        engine.setProperty('pitch', 75)                                 # espeak supports pitch
    except Exception:
        pass

    @staticmethod
    def speak(*texts):
        """Speak one or more text snippets safely."""
        for t in texts:
            GUI_Exam.engine.say(t)
        try:
            GUI_Exam.engine.runAndWait()
        except RuntimeError:
            GUI_Exam.engine.stop()
            GUI_Exam.engine.runAndWait()

    root = Tk()
    
    @classmethod
    def launch_main(cls):
        """
        Generate and display a new math question.
        """
        cls.root.title("MathQuest Adventures")
        cls.root.state('zoomed')
        cls.root.geometry("1530x775")
        icon_path = resource_path("Icon.ico")
        if os.path.exists(icon_path):
            try:
                cls.root.iconbitmap(icon_path)
            except Exception:
                pass
        cls.root.configure(bg="#F0F8FF")
        return cls()
            
    def __init__(self):
        """
        Initialize the GUI_Exam object.

        Parameters:
            - master: Parent Tkinter window (default is None)
        """
        self.bg_color = "#F0F8FF"
        self.home_frame = Frame(GUI_Exam.root, bg=self.bg_color)
        # canvas + scrollbar for scrollable home screen
        self.home_canvas = Canvas(self.home_frame, bg=self.bg_color, highlightthickness=0)
        self.home_scrollbar = Scrollbar(self.home_frame, orient="vertical", command=self.home_canvas.yview)
        self.home_canvas.configure(yscrollcommand=self.home_scrollbar.set)
        self.home_scroll_frame = Frame(self.home_canvas, bg=self.bg_color)
        self.scroll_window = self.home_canvas.create_window(
            (0, 0), window=self.home_scroll_frame, anchor="n"
        )
        self.home_scroll_frame.bind(
            "<Configure>",
            lambda e: self._update_scrollregion(),
        )
        self.home_canvas.bind(
            "<Configure>",
            lambda e: self.home_canvas.coords(self.scroll_window, e.width / 2, 0),
        )
        self.container = Frame(self.home_scroll_frame, bg=self.bg_color)
        self.banner_label = Label(self.container, text="🧮", font=("Comic Sans MS", 60), bg=self.bg_color)
        self.home_label = Label(
            self.container,
            text="Welcome to the MathQuest Adventures",
            font=("Comic Sans MS", 40, "bold"),
            justify="center",
            bg=self.bg_color,
        )
        self.aritmatic_label = Label(
            self.container,
            text="Please select Arithmatics:",
            font=("Comic Sans MS", 24),
            justify="left",
            bg=self.bg_color,
        )
        self.add_variable = StringVar()
        self.subtract_variable = StringVar()
        self.multiply_variable = StringVar()
        self.divide_variable = StringVar()
        self.fraction_variable = StringVar()
        self.factors_primes_variable = StringVar()
        self.prime_factor_variable = StringVar()
        self.hcf_variable = StringVar()
        self.lcm_variable = StringVar()
        self.select_all_variable = StringVar()
        self.display_question = StringVar()
        self.grade = StringVar()
        self.sound_variable = StringVar()

        self.basic_ops_frame = LabelFrame(
            self.container,
            text="Arithmetic",
            font=("Comic Sans MS", 18),
            bg=self.bg_color,
            padx=10,
            pady=10,
        )
        self.adv_ops_frame = LabelFrame(
            self.container,
            text="More Options",
            font=("Comic Sans MS", 18),
            bg=self.bg_color,
            padx=10,
            pady=10,
        )

        cbfont = ("Comic Sans MS", 18)
        self.add_checkbox = Checkbutton(
            self.basic_ops_frame,
            text="Addition",
            variable=self.add_variable,
            onvalue="+",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=15,
            bg=self.bg_color,
        )
        self.subtract_checkbox = Checkbutton(
            self.basic_ops_frame,
            text="Subraction",
            variable=self.subtract_variable,
            onvalue="-",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=15,
            bg=self.bg_color,
        )
        self.multiply_checkbox = Checkbutton(
            self.basic_ops_frame,
            text="Multiplication",
            variable=self.multiply_variable,
            onvalue="*",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=15,
            bg=self.bg_color,
        )
        self.divide_checkbox = Checkbutton(
            self.basic_ops_frame,
            text="Division",
            variable=self.divide_variable,
            onvalue="/",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=15,
            bg=self.bg_color,
        )
        self.fraction_checkbox = Checkbutton(
            self.adv_ops_frame,
            text="Fractions",
            variable=self.fraction_variable,
            onvalue="fraction",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=18,
            bg=self.bg_color,
        )
        self.factors_primes_checkbox = Checkbutton(
            self.adv_ops_frame,
            text="Factors & Primes",
            variable=self.factors_primes_variable,
            onvalue="factors_primes",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=18,
            bg=self.bg_color,
        )
        self.prime_factor_checkbox = Checkbutton(
            self.adv_ops_frame,
            text="Prime Factorization",
            variable=self.prime_factor_variable,
            onvalue="prime_factorization",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=18,
            bg=self.bg_color,
        )
        self.hcf_checkbox = Checkbutton(
            self.adv_ops_frame,
            text="HCF",
            variable=self.hcf_variable,
            onvalue="hcf",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=18,
            bg=self.bg_color,
        )
        self.lcm_checkbox = Checkbutton(
            self.adv_ops_frame,
            text="LCM",
            variable=self.lcm_variable,
            onvalue="lcm",
            offvalue=None,
            font=cbfont,
            anchor="w",
            width=18,
            bg=self.bg_color,
        )
        self.select_all_checkbox = Checkbutton(
            self.container,
            text="All of the above!",
            variable=self.select_all_variable,
            onvalue="select_all",
            offvalue=None,
            font=("Comic Sans MS", 18),
            bg=self.bg_color,
            anchor="w",
        )
        self.add_checkbox.deselect(), self.subtract_checkbox.deselect(), self.multiply_checkbox.deselect()
        self.divide_checkbox.deselect(), self.fraction_checkbox.deselect(), self.factors_primes_checkbox.deselect(), self.prime_factor_checkbox.deselect(), self.hcf_checkbox.deselect(), self.lcm_checkbox.deselect(), self.select_all_checkbox.deselect()
        self.label_num_question = Label(
            self.container,
            text="Type number of Questions:",
            font=("Comic Sans MS", 20),
            justify="left",
            bg=self.bg_color,
        )
        self.input_num_question = Entry(
            self.container,
            font=("Comic Sans MS", 20),
            justify="center",
            width=4,
        )
        self.start_exam_button = Button(
            self.container,
            text="Start Exam!",
            font=("Comic Sans MS", 16),
            command=self.start,
        )
        self.progress_button = Button(
            self.container,
            text="\U0001F4CA View Progress",
            font=("Comic Sans MS", 16),
            command=self.launch_progress_dashboard,
        )
        self.test_checkbox = Label(
            self.container,
            text="Please ensure correct selections & entry!",
            font=("Times", 20),
            bg="red",
        )
        self.exam_frame = Frame(GUI_Exam.root, bg=self.bg_color)
        self.icon_label = Label(
            self.exam_frame,
            text="🧮",
            font=("Comic Sans MS", 50),
            bg=self.bg_color,
        )
        self.question_box = LabelFrame(
            self.exam_frame,
            bg=self.bg_color,
            bd=4,
            relief="groove",
        )
        self.status_checkbox, self.question_to_ask = None, None
        self.question_label = Label(
            self.question_box,
            text=self.display_question.get(),
            font=("Comic Sans MS", 32, "bold"),
            bg=self.bg_color,
            justify="center",
            wraplength=1000,
        )
        self.label_user_answer = Label(
            self.exam_frame,
            text="Type Answer Here:",
            font=("Comic Sans MS", 24),
            bg=self.bg_color,
            justify="center",
        )
        self.input_user_answer = Entry(
            self.exam_frame,
            font=("Comic Sans MS", 24),
            justify="center",
            width=7,
        )
        self.label_user_answer_remainder = Label(
            self.exam_frame,
            text="Type Remainder Here:",
            font=("Comic Sans MS", 24),
            bg=self.bg_color,
            justify="center",
        )
        self.input_user_answer_remainder = Entry(
            self.exam_frame,
            font=("Comic Sans MS", 24),
            justify="center",
            width=7,
        )
        self.choice_var = IntVar()
        self.options_frame = None
        self.question_asked, self.exam_score = 0, 0          # To keep track of the number of questions & correct answers.
        self.start_time, self.test_start, self.question_paper = None, None, None
        self.attempts_counter = 0
        self.check_button = Button(
            self.exam_frame,
            text="Submit",
            font=("Comic Sans MS", 20),
            command=self.check_user_answer,
        )
        self.evaluation_feedback = Label(
            self.exam_frame,
            font=("Comic Sans MS", 24),
            bg=self.bg_color,
            justify="center",
            wraplength=900,
        )
        self.evaluation_result, self.end_time, self.test_end = None, None, None
        self.result_frame = Frame(self.root)
        self.grade_label = Label(self.result_frame, font=("Bell MT", 50), justify="center", width=38)
        self.stat_frame = Frame(self.result_frame, width=350, height=475, bd=5, relief="groove")
        self.quit_button = Button(self.result_frame, text="Quit!", font=("Bell MT", 16), command=self.root.quit)
        self.sound_checkbox = Checkbutton(
            self.exam_frame,
            text="Disable Sound!",
            variable=self.sound_variable,
            onvalue="",
            offvalue="Enable",
            font=("Comic Sans MS", 20),
            bd=5,
            relief="groove",
            bg=self.bg_color,
        )
        self.file_name = f"Practice_dated_{datetime.now().strftime('%d-%b-%y-%I%M')}"
        self.file_open_mode = None
        self.pdf = None
        self.stats = {}
        self.launch_home_frame()
                
    def launch_home_frame(self):
        self.home_frame.pack(fill="both", expand=1)
        # pack scrollable canvas and scrollbar
        self.home_canvas.pack(side="left", fill="both", expand=True)
        self.home_scrollbar.pack(side="right", fill="y")
        self.container.pack(expand=True)
        self.home_canvas.yview_moveto(0)
        self.home_frame.update_idletasks()
        self._update_scrollregion()

        # banner and title
        self.banner_label.grid(row=0, column=0, columnspan=2, pady=(20, 10))
        self.home_label.grid(row=1, column=0, columnspan=2, pady=(0, 20))

        self.aritmatic_label.grid(row=2, column=0, columnspan=2, pady=(0, 10), sticky="w")

        # operation frames
        self.basic_ops_frame.grid(row=3, column=0, padx=10, sticky="n")
        self.adv_ops_frame.grid(row=3, column=1, padx=10, sticky="n")

        # pack checkboxes inside frames
        for widget in (
            self.add_checkbox,
            self.subtract_checkbox,
            self.multiply_checkbox,
            self.divide_checkbox,
        ):
            widget.pack(anchor="w")

        for widget in (
            self.fraction_checkbox,
            self.factors_primes_checkbox,
            self.prime_factor_checkbox,
            self.hcf_checkbox,
            self.lcm_checkbox,
        ):
            widget.pack(anchor="w")

        self.select_all_checkbox.grid(row=4, column=0, columnspan=2, pady=(10, 10), sticky="w")
        self.label_num_question.grid(row=5, column=0, sticky="e")
        self.input_num_question.grid(row=5, column=1, sticky="w")
        self.start_exam_button.grid(row=6, column=0, columnspan=2, pady=(20, 5))
        self.progress_button.grid(row=7, column=0, columnspan=2, pady=(5, 20))

        # bind mousewheel scrolling for canvas
        self.home_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.home_canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.home_canvas.bind_all("<Button-5>", self._on_mousewheel)

    def _on_mousewheel(self, event):
        """Scroll the canvas vertically on mouse wheel events."""
        if event.delta:
            self.home_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        elif event.num == 4:
            self.home_canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.home_canvas.yview_scroll(1, "units")

    def _update_scrollregion(self):
        """Update canvas scrollregion and toggle scrollbar."""
        self.home_canvas.configure(scrollregion=self.home_canvas.bbox("all"))
        self.home_canvas.coords(
            self.scroll_window, self.home_canvas.winfo_width() / 2, 0
        )
        region = self.home_canvas.bbox("all")
        if region and region[3] <= self.home_canvas.winfo_height():
            if self.home_scrollbar.winfo_ismapped():
                self.home_scrollbar.pack_forget()
        else:
            if not self.home_scrollbar.winfo_ismapped():
                self.home_scrollbar.pack(side="right", fill="y")

    def checkbox_status(self):
        if self.select_all_variable.get() == "select_all" and not self.input_num_question.get() == "" and str(self.input_num_question.get()).isdecimal() and int(self.input_num_question.get()) > 0:
            self.add_variable.set("+")
            self.subtract_variable.set("-")
            self.multiply_variable.set("*")
            self.divide_variable.set("/")
            self.fraction_variable.set("fraction")
            self.factors_primes_variable.set("factors_primes")
            self.prime_factor_variable.set("prime_factorization")
            self.hcf_variable.set("hcf")
            self.lcm_variable.set("lcm")
        status_list = [
            self.add_variable.get(),
            self.subtract_variable.get(),
            self.multiply_variable.get(),
            self.divide_variable.get(),
            self.fraction_variable.get(),
            self.factors_primes_variable.get(),
            self.prime_factor_variable.get(),
            self.hcf_variable.get(),
            self.lcm_variable.get(),
        ]
        if all(item in ("0", "", None) for item in status_list):
            return "Please Select atleast One option!"
        else:
            return [
                self.add_variable.get(),
                self.subtract_variable.get(),
                self.multiply_variable.get(),
                self.divide_variable.get(),
                self.fraction_variable.get(),
                self.factors_primes_variable.get(),
                self.prime_factor_variable.get(),
                self.hcf_variable.get(),
                self.lcm_variable.get(),
            ]
    
    def start(self):
        """Start the exam based on user selections."""
        self.test_checkbox.grid_forget()
        if self.checkbox_status() == "Please Select atleast One option!" or self.input_num_question.get() == "" or not str(self.input_num_question.get()).isdecimal() or int(self.input_num_question.get()) <= 0:
            self.test_checkbox.grid(row=8, column=0, columnspan=2, pady=(5, 0))
        else:
            self.launch_exam_frame()

    def launch_factor_mode(self):
        self.home_frame.pack_forget()
        self.home_canvas.unbind_all("<MouseWheel>")
        self.home_canvas.unbind_all("<Button-4>")
        self.home_canvas.unbind_all("<Button-5>")
        self.factor_frame = Frame(GUI_Exam.root)
        self.factor_frame.pack(fill="both", expand=1)
        self.factor_count = 0
        Label(self.factor_frame, text="Enter a number between 2 and 100:", font=("Bell MT", 30)).grid(row=0, column=0, columnspan=3, pady=20)
        self.factor_entry = Entry(self.factor_frame, font=("Bell MT", 20), justify="center", width=7)
        self.factor_entry.grid(row=1, column=0, columnspan=3)
        self.factor_submit = Button(self.factor_frame, text="Submit", font=("Bell MT", 16), command=self.process_factor_input)
        self.factor_submit.grid(row=2, column=0, columnspan=3, pady=10)
        self.factor_feedback = Label(self.factor_frame, font=("Bell MT", 20), wraplength=1000, justify="left")
        self.factor_feedback.grid(row=3, column=0, columnspan=3, pady=20)
        self.factor_back_button = Button(self.factor_frame, text="Back to Home", font=("Bell MT", 14), command=self.back_from_factor)

    def process_factor_input(self):
        val = self.factor_entry.get()
        if not val.isdecimal():
            messagebox.showerror("Input Error", "Please type Numbers only!")
            if self.sound_variable.get() != "":
                GUI_Exam.speak(self.for_incorrect_answer())
            return
        n = int(val)
        if n < 2 or n > 100:
            self.factor_feedback.config(text="Please enter a number between 2 and 100", bg="yellow")
            if self.sound_variable.get() != "":
                GUI_Exam.speak(self.for_incorrect_answer())
            return
        facs = factors_of(n)
        if len(facs) == 2:
            status = "a prime number"
        else:
            status = "a composite number" if n != 1 else "neither prime nor composite"
        pair = twin_prime_pair(n)
        pair_text = f" and part of the twin prime pair {pair}" if pair else ""
        msg = f"Factors of {n}: {', '.join(map(str, facs))}. It is {status}{pair_text}."
        self.factor_feedback.config(text=msg, bg="lightgreen")
        if self.sound_variable.get() != "":
            GUI_Exam.speak(self.for_correct_answer(), msg)
        self.factor_count += 1
        if self.factor_count >= 3:
            self.factor_submit.config(state=DISABLED)
            self.factor_back_button.grid(row=4, column=0, columnspan=3)
        self.factor_entry.delete(0, END)

    def back_from_factor(self):
        self.factor_frame.pack_forget()
        self.launch_home_frame()

    def launch_progress_dashboard(self):
        """Open a window showing progress charts from AllSessions.xlsx."""
        path = os.path.join(OUTPUT_DIR, "AllSessions.xlsx")
        if not os.path.exists(path):
            messagebox.showinfo("Progress", "No session data found yet.")
            return

        try:
            log_df = pd.read_excel(path, sheet_name="Log")
            idx_df = pd.read_excel(path, sheet_name="Index")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")
            return

        # The log sheet stores start times with a date component while the
        # index only contains the time. Extract the time portion before
        # merging so we can match rows correctly.
        log_df["_time"] = log_df["Start Time"].astype(str).str[-8:]
        idx_df["_time"] = idx_df["Start Time"].astype(str).str[-8:]
        log_df = log_df.merge(idx_df[["_time", "Session Number"]], on="_time", how="left")
        log_df.drop(columns=["_time"], inplace=True)

        # Use forward fill for the session numbers so consecutive rows share
        # the appropriate session identifier.
        log_df["Session Number"] = log_df["Session Number"].ffill()

        diff_hist = load_difficulty_history()

        dash = Toplevel(self.root)
        dash.title("Progress Dashboard")
        dash.configure(bg=self.bg_color)

        for i in range(3):
            dash.rowconfigure(i, weight=1)
            dash.columnconfigure(i, weight=1)

        frames = {}
        for r in range(3):
            for c in range(3):
                f = Frame(dash, bg=self.bg_color)
                f.grid(row=r, column=c, sticky="nsew", padx=5, pady=5)
                frames[(r, c)] = f

        def add_chart(draw_func, row, col, colspan=1):
            container = frames[(row, col)]
            if colspan > 1:
                container.grid(columnspan=colspan)
            fig = Figure(figsize=(3, 2), dpi=100)
            ax = fig.add_subplot(111)
            draw_func(ax)
            fig.tight_layout()
            canvas = FigureCanvasTkAgg(fig, master=container)
            canvas.draw()
            widget = canvas.get_tk_widget()
            widget.pack(fill="both", expand=True)

            def open_large(event=None, func=draw_func):
                top = Toplevel(dash)
                top.title("Chart")
                fig2 = Figure(figsize=(6, 4), dpi=100)
                ax2 = fig2.add_subplot(111)
                func(ax2)
                fig2.tight_layout()
                canv2 = FigureCanvasTkAgg(fig2, master=top)
                canv2.draw()
                canv2.get_tk_widget().pack(fill="both", expand=True)

            widget.bind("<Button-1>", open_large)

        def accuracy_over_time(ax):
            data = (
                log_df.groupby(["Session Number", "Question Type"])["Accuracy (%)"].mean().reset_index()
            )
            pivot = data.pivot(index="Session Number", columns="Question Type", values="Accuracy (%)")
            pivot.plot(ax=ax, marker="o")
            ax.set_title("Accuracy Over Time")
            ax.set_xlabel("Session")
            ax.set_ylabel("Accuracy (%)")
            ax.legend(loc="best", fontsize="x-small")

        def difficulty_evolution(ax):
            if not diff_hist:
                ax.text(0.5, 0.5, "No data", ha="center", va="center")
                ax.set_axis_off()
                return
            max_len = max(len(v) for v in diff_hist.values())
            df = pd.DataFrame({op_names.get(k, k): v + [None] * (max_len - len(v)) for k, v in diff_hist.items()})
            df.index = range(1, max_len + 1)
            df.plot(ax=ax, marker="o")
            ax.set_title("Difficulty Score Evolution")
            ax.set_xlabel("Session")
            ax.set_ylabel("Difficulty Score")
            ax.legend(loc="best", fontsize="x-small")

        def session_score(ax):
            idx_df.plot(x="Session Number", y="Accuracy (%)", kind="bar", ax=ax)
            ax.set_title("Session Score Trend")
            ax.set_xlabel("Session")
            ax.set_ylabel("Accuracy (%)")
            if ax.legend_:
                ax.legend_.remove()

        def topic_accuracy(ax):
            data = log_df.groupby("Question Type").apply(
                lambda g: g["Correct Answers"].sum() / g["Total Questions"].sum() * 100
            )
            data.plot(kind="bar", ax=ax)
            ax.set_title("Topic-wise Accuracy")
            ax.set_xlabel("Operation")
            ax.set_ylabel("Accuracy (%)")

        def time_of_day_accuracy(ax):
            hrs = pd.to_datetime(idx_df["Start Time"], errors="coerce").dt.hour
            ax.scatter(hrs, idx_df["Accuracy (%)"])
            ax.set_title("Time of Day vs Accuracy")
            ax.set_xlabel("Hour of Day")
            ax.set_ylabel("Accuracy (%)")
            ax.set_xticks(range(0, 24, 1))
            ax.set_xticklabels(range(0, 24, 1), rotation=45)

        def duration_vs_accuracy(ax):
            ax.scatter(idx_df["Duration"], idx_df["Accuracy (%)"])
            ax.set_title("Duration vs Accuracy")
            ax.set_xlabel("Duration (min)")
            ax.set_ylabel("Accuracy (%)")

        def topic_distribution(ax):
            dist = log_df.groupby("Question Type")["Total Questions"].sum()
            dist.plot(kind="pie", ax=ax, autopct="%1.0f%%")
            ax.set_title("Topic Distribution")
            ax.set_ylabel("")

        def difficulty_vs_accuracy(ax):
            if not diff_hist:
                ax.text(0.5, 0.5, "No data", ha="center", va="center")
                ax.set_axis_off()
                return
            rows = []
            for op, vals in diff_hist.items():
                for i, val in enumerate(vals, start=1):
                    rows.append(
                        {
                            "Session Number": i,
                            "Question Type": op_names.get(op, op),
                            "Difficulty Score": val,
                        }
                    )
            diff_df = pd.DataFrame(rows)
            merged = log_df.merge(diff_df, on=["Session Number", "Question Type"], how="inner")
            ax.scatter(
                merged["Accuracy (%)"],
                merged["Difficulty Score"],
                s=merged["Total Questions"] * 5,
                alpha=0.6,
            )
            ax.set_title("Difficulty Score vs Accuracy")
            ax.set_xlabel("Accuracy (%)")
            ax.set_ylabel("Avg Difficulty Score")

        def attempts_per_type(ax):
            data = log_df.groupby("Question Type").apply(
                lambda g: g["Total Attempts"].sum() / g["Total Questions"].sum()
            )
            data.plot(kind="bar", ax=ax)
            ax.set_title("Attempts per Question Type")
            ax.set_xlabel("Operation")
            ax.set_ylabel("Avg Attempts")

        add_chart(accuracy_over_time, 0, 0)
        add_chart(difficulty_evolution, 0, 1)
        add_chart(session_score, 0, 2)

        add_chart(topic_accuracy, 1, 0)
        add_chart(topic_distribution, 1, 1)
        add_chart(duration_vs_accuracy, 1, 2)

        add_chart(time_of_day_accuracy, 2, 0)
        add_chart(difficulty_vs_accuracy, 2, 1)
        add_chart(attempts_per_type, 2, 2)


    def launch_exam_frame(self):
        self.status_checkbox = self.checkbox_status()                 # To fetch the user selection
        self.question_to_ask = int(self.input_num_question.get())     # To fetch how many question to ask
        self.stats = {s: {"total_questions": 0, "correct_answers": 0, "total_attempts": 0,
                          "total_time": 0.0, "first_try_correct": 0}
                       for s in self.status_checkbox if s not in (None, "", "0")}
        self.prepare_question_plan()
        self.home_frame.pack_forget()
        self.home_canvas.unbind_all("<MouseWheel>")
        self.home_canvas.unbind_all("<Button-4>")
        self.home_canvas.unbind_all("<Button-5>")
        self.exam_frame.pack(fill="both", expand=1)
        self.sound_checkbox.grid(row=0, column=7, pady=10, sticky="e")
        self.sound_checkbox.deselect()
        self.icon_label.grid(row=1, column=0, columnspan=8, pady=(10, 5))
        self.question_box.grid(row=2, column=1, columnspan=6, pady=10, padx=10)
        self.question_label.pack(padx=20, pady=20)
        self.label_user_answer.grid(row=6, column=1, sticky="E", pady=10)
        self.input_user_answer.grid(row=6, column=2, sticky="W", pady=10, padx=5)
        self.start_time = datetime.now()
        self.test_start = self.start_time.strftime("%I:%M%p")
        self.check_button.grid(row=10, column=1, columnspan=2, pady=10)
        self.generate_question()

    def prepare_question_plan(self):
        """Build a plan of operations and difficulty levels for this session."""
        history = load_difficulty_history()
        stats = compute_thresholds(history)
        self.levels = determine_difficulty_levels(difficulty_scores, stats)

        total = self.question_to_ask
        base = total // 3
        dist = {"Easy": base, "Medium": base, "Hard": base}
        for i in range(total - base * 3):
            dist[["Easy", "Medium", "Hard"][i]] += 1

        ops_by_level = {"Easy": [], "Medium": [], "Hard": []}
        for op in self.status_checkbox:
            ops_by_level[self.levels.get(op, "Medium")].append(op)

        plan = []
        for lvl in ["Easy", "Medium", "Hard"]:
            ops = ops_by_level[lvl] or self.status_checkbox
            for _ in range(dist[lvl]):
                plan.append((random.choice(ops), lvl))
        random.shuffle(plan)
        self.question_plan = plan
        self.question_index = 0
        
    def generate_question(self):
        """
        Generate and display a new math question.
        """
        op, level = self.question_plan[self.question_index]
        self.question_index += 1
        self.question_paper = Exam.quiz(op, level)
        if self.question_paper._S not in self.stats:
            self.stats[self.question_paper._S] = {"total_questions": 0, "correct_answers": 0,
                                                "total_attempts": 0, "total_time": 0.0,
                                                "first_try_correct": 0}
        self.stats[self.question_paper._S]["total_questions"] += 1
        if self.question_paper._S in ["+", "-", "*", "/"]:
            formatted = (
                f"Q.{self.question_asked + 1} What will be the result of {self.question_paper.question}?"
            )
        else:
            formatted = f"Q.{self.question_asked + 1} {self.question_paper.question}"
        self.display_question.set(formatted)
        self.question_label.config(text=formatted)
        self.current_question_start = datetime.now()
        self.question_asked += 1

        if self.options_frame:
            self.options_frame.destroy()
            self.options_frame = None
        self.choice_var.set(-1)

        if self.question_paper._S == "/":
            self.label_user_answer.config(text="Type Quotient Here:")
            self.label_user_answer_remainder.grid(
                row=6,
                column=3,
                rowspan=2,
                columnspan=1,
                sticky="E",
            )
            self.input_user_answer_remainder.grid(
                row=6,
                column=4,
                rowspan=2,
                columnspan=1,
                sticky="W",
            )
            self.label_user_answer.grid(
                row=6,
                column=1,
                rowspan=2,
                columnspan=1,
                sticky="E",
            )
            self.input_user_answer.grid(
                row=6,
                column=2,
                rowspan=2,
                columnspan=1,
                sticky="W",
            )
        elif self.question_paper._S == "fraction":
            if self.question_paper.choices:
                self.label_user_answer.grid_forget()
                self.input_user_answer.grid_forget()
                self.label_user_answer_remainder.grid_forget()
                self.input_user_answer_remainder.grid_forget()
                self.options_frame = Frame(self.exam_frame)
                self.options_frame.grid(row=6, column=1, columnspan=6)
                for i, frac in enumerate(self.question_paper.choices):
                    frame = Frame(self.options_frame)
                    # Increase the option size for better visibility
                    canvas_width, canvas_height = 120, 80
                    canvas = Canvas(frame, width=canvas_width, height=canvas_height)
                    for j in range(frac[1]):
                        x0 = j * (canvas_width / frac[1])
                        x1 = (j + 1) * (canvas_width / frac[1])
                        color = "blue" if j < frac[0] else "white"
                        canvas.create_rectangle(x0, 0, x1, canvas_height, fill=color, outline="black")
                    canvas.pack()
                    Radiobutton(
                        frame,
                        variable=self.choice_var,
                        value=i,
                        font=("Comic Sans MS", 18),
                        bg=self.bg_color,
                    ).pack(pady=5)
                    frame.grid(row=0, column=i, padx=5)
            else:
                self.label_user_answer.config(text="Type Answer Here:")
                self.label_user_answer_remainder.grid_forget()
                self.input_user_answer_remainder.grid_forget()
                self.label_user_answer.grid(
                    row=6,
                    column=1,
                    rowspan=2,
                    columnspan=1,
                    sticky="E",
                )
                self.input_user_answer.grid(
                    row=6,
                    column=2,
                    rowspan=2,
                    columnspan=1,
                    sticky="W",
                )
        elif self.question_paper._S == "prime_factorization":
            self.label_user_answer.config(text="Enter prime factors:")
            self.label_user_answer_remainder.grid_forget()
            self.input_user_answer_remainder.grid_forget()
            self.label_user_answer.grid(row=6, column=1, rowspan=2, columnspan=1, sticky="E")
            self.input_user_answer.grid(row=6, column=2, rowspan=2, columnspan=1, sticky="W")
        else:
            self.label_user_answer.config(text="Type Answer Here:")
            self.label_user_answer_remainder.grid_forget()
            self.input_user_answer_remainder.grid_forget()
            self.label_user_answer.grid(
                row=6,
                column=1,
                rowspan=2,
                columnspan=1,
                sticky="E",
            )
            self.input_user_answer.grid(
                row=6,
                column=2,
                rowspan=2,
                columnspan=1,
                sticky="W",
            )
    
    def check_user_answer(self):
        """
        Check the user's answer and provide feedback.
        """
        if self.question_paper._S == "/":
            if self.input_user_answer.get().isdecimal() and self.input_user_answer_remainder.get().isdecimal():
                self.question_paper.answer_user = int(self.input_user_answer.get())
                self.question_paper.answer_user_remainder = int(self.input_user_answer_remainder.get())
            else:
                messagebox.showerror("Input Error", "Please type Numbers only!")
                return
        elif self.question_paper._S == "fraction":
            if self.question_paper.choices:
                if self.choice_var.get() == -1:
                    messagebox.showerror("Input Error", "Please select an option!")
                    return
                self.question_paper.answer_user = self.choice_var.get()
            else:
                parsed = parse_fraction_input(self.input_user_answer.get())
                if parsed is None:
                    messagebox.showerror(
                        "Input Error",
                        "Please enter a number or fraction like a/b",
                    )
                    return
                self.question_paper.answer_user = parsed
        elif self.question_paper._S == "factors_primes":
            if self.input_user_answer.get().isdecimal():
                self.question_paper.answer_user = int(self.input_user_answer.get())
            else:
                messagebox.showerror("Input Error", "Please type Numbers only!")
                return
        elif self.question_paper._S == "prime_factorization":
            parsed = parse_factor_input(self.input_user_answer.get())
            if parsed is None:
                messagebox.showerror(
                    "Input Error",
                    "Please enter prime factors separated by ×, * or spaces",
                )
                return
            self.question_paper.answer_user = parsed
        elif self.question_paper._S == "hcf":
            if self.input_user_answer.get().isdecimal():
                self.question_paper.answer_user = int(self.input_user_answer.get())
            else:
                messagebox.showerror("Input Error", "Please type Numbers only!")
                return
        else:
            if self.input_user_answer.get().isdecimal():
                self.question_paper.answer_user = int(self.input_user_answer.get())
            else:
                messagebox.showerror("Input Error", "Please type Numbers only!")
                return

        self.stats[self.question_paper._S]["total_attempts"] += 1

        if self.question_paper._S == "factors_primes":
            self.evaluation_result = (
                self.question_paper.answer_user == self.question_paper.answer_actual
            )
        elif self.question_paper._S == "prime_factorization":
            user = Counter(self.question_paper.answer_user)
            actual = Counter(self.question_paper.answer_actual)
            self.evaluation_result = (
                user == actual and all(is_prime(f) for f in self.question_paper.answer_user)
            )
        elif self.question_paper._S == "fraction" and not self.question_paper.choices:
            actual = self.question_paper.answer_actual
            user = self.question_paper.answer_user
            if isinstance(actual, tuple):
                self.evaluation_result = actual[0] * user[1] == user[0] * actual[1]
            else:
                self.evaluation_result = abs(actual - (user[0] / user[1])) < 1e-6
        else:
            self.evaluation_result = evaluate(
                self.question_paper.answer_user,
                self.question_paper.answer_actual,
                self.question_paper.answer_user_remainder,
                self.question_paper.answer_actual_remainder if self.question_paper._S == "/" else None,
                self.question_paper._S,
            )
        if self.evaluation_result == True:
            if self.question_paper._S == "/":
                self.evaluation_feedback.config(
                    text=f"Correct!, For {self.question_paper.question}, the Quotient is {self.question_paper.answer_actual} & Remainder is {self.question_paper.answer_actual_remainder}",
                    bg="green",
                )
            elif self.question_paper._S == "fraction":
                if self.question_paper.choices:
                    self.evaluation_feedback.config(
                        text="Correct!",
                        bg="green",
                    )
                else:
                    ans = (
                        f"{self.question_paper.answer_actual[0]}/{self.question_paper.answer_actual[1]}"
                        if isinstance(self.question_paper.answer_actual, tuple)
                        else str(self.question_paper.answer_actual)
                    )
                    self.evaluation_feedback.config(
                        text=f"Correct!, {self.question_paper.question} is {ans}",
                        bg="green",
                    )
            elif self.question_paper._S == "factors_primes":
                number = self.question_paper._X
                facs = self.question_paper.factors
                status = (
                    "a prime number"
                    if len(facs) == 2
                    else (
                        "neither prime nor composite"
                        if number == 1
                        else "a composite number"
                    )
                )
                pair = self.question_paper.twin_pair
                pair_text = f" and part of the twin prime pair {pair}" if pair else ""
                explanation = f"Factors of {number}: {', '.join(map(str, facs))}. It is {status}{pair_text}."
                self.evaluation_feedback.config(
                    text=f"Correct! {explanation}",
                    bg="green",
                )
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_correct_answer(), explanation)
            elif self.question_paper._S == "prime_factorization":
                ans = " × ".join(map(str, sorted(self.question_paper.answer_actual)))
                msg = f"Correct! Prime factorization of {self.question_paper._X} is {ans}"
                self.evaluation_feedback.config(text=msg, bg="green")
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_correct_answer(), msg)
            elif self.question_paper._S == "hcf":
                nums = self.question_paper.numbers
                if len(nums) == 3:
                    ntext = f"{nums[0]}, {nums[1]}, and {nums[2]}"
                else:
                    ntext = f"{nums[0]} and {nums[1]}"
                msg = f"Correct! The HCF of {ntext} is {self.question_paper.answer_actual}."
                self.evaluation_feedback.config(text=msg, bg="green")
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_correct_answer(), msg)
            elif self.question_paper._S == "lcm":
                nums = self.question_paper.numbers
                if len(nums) == 3:
                    ntext = f"{nums[0]}, {nums[1]}, and {nums[2]}"
                else:
                    ntext = f"{nums[0]} and {nums[1]}"
                explanation = lcm_explanation(nums, self.question_paper.method)
                msg = f"Correct! The LCM of {ntext} is {self.question_paper.answer_actual}."
                self.evaluation_feedback.config(text=f"{msg}\n{explanation}", bg="green")
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_correct_answer(), msg, explanation)
            else:
                self.evaluation_feedback.config(
                    text=f"Correct!, {self.question_paper.question} is {self.question_paper.answer_actual}",
                    bg="green",
                )
            self.evaluation_feedback.grid(row=12, column=1, columnspan=8, pady=10)
            self.stats[self.question_paper._S]["correct_answers"] += 1
            self.exam_score += 1
            if self.sound_variable.get() != "":
                if self.question_paper._S not in ["factors_primes", "prime_factorization"]:
                    GUI_Exam.speak(self.for_correct_answer())
        else:
            if self.attempts_counter == 0:
                self.evaluation_feedback.config(
                    text="Your Answer is Incorrect, you've got 2 more attempts!",
                    bg="teal",
                )
                self.evaluation_feedback.grid(row=12, column=1, columnspan=8, pady=10)
                self.attempts_counter += 1
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_incorrect_answer())
            elif self.attempts_counter == 1:
                self.evaluation_feedback.grid_forget()
                self.evaluation_feedback.config(
                    text="Your Answer is Incorrect, it's the last attempt!",
                    bg="yellow",
                )
                self.evaluation_feedback.grid(row=12, column=1, columnspan=8, pady=10)
                self.attempts_counter += 1
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_incorrect_answer())
            elif self.attempts_counter == 2:
                self.evaluation_feedback.grid_forget()
                if self.question_paper._S == "/":
                    self.evaluation_feedback.config(
                        text=f"Incorrect!, For {self.question_paper.question} the Quotient is {self.question_paper.answer_actual} & Remainder is {self.question_paper.answer_actual_remainder} not {self.question_paper.answer_user} & {self.question_paper.answer_user_remainder}",
                        bg="red",
                    )
                    if self.sound_variable.get() != "":
                        GUI_Exam.speak(
                            f"Incorrect!, For {self.question_paper.question} the Quotient is {self.question_paper.answer_actual} & Remainder is {self.question_paper.answer_actual_remainder}"
                        )
                else:
                    if self.question_paper._S == "fraction":
                        if self.question_paper.choices:
                            self.evaluation_feedback.config(
                                text="Incorrect!",
                                bg="red",
                            )
                            if self.sound_variable.get() != "":
                                GUI_Exam.speak("Incorrect!")
                        else:
                            ans = (
                                f"{self.question_paper.answer_actual[0]}/{self.question_paper.answer_actual[1]}"
                                if isinstance(self.question_paper.answer_actual, tuple)
                                else str(self.question_paper.answer_actual)
                            )
                            user_ans = (
                                f"{self.question_paper.answer_user[0]}/{self.question_paper.answer_user[1]}"
                                if isinstance(self.question_paper.answer_user, tuple)
                                else str(self.question_paper.answer_user)
                            )
                            self.evaluation_feedback.config(
                                text=f"Incorrect!, {self.question_paper.question} is {ans} not {user_ans}",
                                bg="red",
                            )
                            if self.sound_variable.get() != "":
                                GUI_Exam.speak(
                                    f"Incorrect!, {self.question_paper.question} is {ans} not {user_ans}"
                                )
                    elif self.question_paper._S == "factors_primes":
                        number = self.question_paper._X
                        facs = self.question_paper.factors
                        status = (
                            "a prime number"
                            if len(facs) == 2
                            else (
                                "neither prime nor composite"
                                if number == 1
                                else "a composite number"
                            )
                        )
                        pair = self.question_paper.twin_pair
                        pair_text = f" and part of the twin prime pair {pair}" if pair else ""
                        explanation = f"The number {number} has {len(facs)} factors: {', '.join(map(str, facs))}. It is {status}{pair_text}."
                        self.evaluation_feedback.config(
                            text=f"Incorrect! {explanation}",
                            bg="red",
                        )
                        if self.sound_variable.get() != "":
                            GUI_Exam.speak(f"Incorrect! {explanation}")
                    elif self.question_paper._S == "prime_factorization":
                        ans = " × ".join(map(str, sorted(self.question_paper.answer_actual)))
                        msg = f"Incorrect. The correct prime factorization of {self.question_paper._X} is {ans}"
                        self.evaluation_feedback.config(text=msg, bg="red")
                        if self.sound_variable.get() != "":
                            GUI_Exam.speak(msg)
                    elif self.question_paper._S == "hcf":
                        nums = self.question_paper.numbers
                        if len(nums) == 3:
                            ntext = f"{nums[0]}, {nums[1]}, and {nums[2]}"
                        else:
                            ntext = f"{nums[0]} and {nums[1]}"
                        msg = f"Incorrect. The correct HCF of {ntext} is {self.question_paper.answer_actual}."
                        self.evaluation_feedback.config(text=msg, bg="red")
                        if self.sound_variable.get() != "":
                            GUI_Exam.speak(msg)
                    elif self.question_paper._S == "lcm":
                        nums = self.question_paper.numbers
                        if len(nums) == 3:
                            ntext = f"{nums[0]}, {nums[1]}, and {nums[2]}"
                        else:
                            ntext = f"{nums[0]} and {nums[1]}"
                        explanation = lcm_explanation(nums, self.question_paper.method)
                        msg = f"Incorrect. The LCM of {ntext} is {self.question_paper.answer_actual}."
                        self.evaluation_feedback.config(text=f"{msg}\n{explanation}", bg="red")
                        if self.sound_variable.get() != "":
                            GUI_Exam.speak(msg, explanation)
                    else:
                        self.evaluation_feedback.config(
                            text=f"Incorrect!, {self.question_paper.question} is {self.question_paper.answer_actual} not {self.question_paper.answer_user}",
                            bg="red",
                        )
                        if self.sound_variable.get() != "":
                            GUI_Exam.speak(
                                f"Incorrect!, {self.question_paper.question} is {self.question_paper.answer_actual} not {self.question_paper.answer_user}"
                            )
                self.evaluation_feedback.grid(row=12, column=1, columnspan=8, pady=10)
                self.attempts_counter += 1
                if self.sound_variable.get() != "":
                    GUI_Exam.speak(self.for_failed_attempt())

        # record time taken for this question
        elapsed = (datetime.now() - self.current_question_start).total_seconds()
        stats = self.stats[self.question_paper._S]
        stats["total_time"] += elapsed
        if self.evaluation_result and self.attempts_counter == 0:
            stats["first_try_correct"] += 1

        # record time taken for this question
        elapsed = (datetime.now() - self.current_question_start).total_seconds()
        stats = self.stats[self.question_paper._S]
        stats["total_time"] += elapsed
        if self.evaluation_result and self.attempts_counter == 0:
            stats["first_try_correct"] += 1

        # Check if all questions have been asked
        if self.question_asked < self.question_to_ask and (self.evaluation_result == True or self.attempts_counter > 2):
            self.store_data()
            self.attempts_counter = 0
            self.generate_question()
        elif self.question_asked <= self.question_to_ask and self.evaluation_result != True and self.attempts_counter <= 2:
            pass
        elif self.question_asked == self.question_to_ask and (self.evaluation_result == True or self.attempts_counter > 2):
            self.store_data()
            self.end_time = datetime.now()
            self.test_end = self.end_time.strftime("%I:%M%p")
            self.launch_result_frame()
        
        # Clear the content of Input Entry Box
        self.input_user_answer.delete(0, END)
        self.input_user_answer_remainder.delete(0, END)

    def launch_result_frame(self):
        """Launch the result screen after completing the exam."""
        self.exam_frame.pack_forget()
        self.result_frame.pack(fill="both", expand=1)
        
        # 3.1 Creating a display for grades
        self.grade.set(get_grade(self.exam_score, self.question_asked))
        self.grade_label.config(text=self.grade.get())
        Label(self.result_frame, width=25, height=3).grid(row=0, column=0, columnspan=5)
        self.grade_label.grid(row=1, column=0, rowspan=3, columnspan=8)

        # 3.2 Creating the exam stat chart
        Label(self.result_frame, width=25, height=10).grid(row=3, column=0, columnspan=5)
        self.stat_frame.grid(row=4, column=0, columnspan=10)
        
        # 3.2.1 Creating test start info
        Label(
        self.stat_frame,
        text=f"Test Date: {self.start_time.strftime('%d-%B-%Y')}\nTest started on: {self.test_start}\nTest ended on: {self.test_end}\nExam Duration: {round((self.end_time - self.start_time).total_seconds()/60, 2)} minutes",
        font=("Bell MT", 16), justify="left"
        ).pack(pady=5)
        Label(
        self.stat_frame,
        text=f"Score: {self.exam_score}\nTotal Questions: {self.question_asked}\nPercent Marks: {round(self.exam_score/self.question_asked*100, 2)}%",
        font=("Bell MT", 16), justify="left"
        ).pack(pady=5)
        
        # 3.2.2 Adding quit button
        Label(self.result_frame, width=25, height=3).grid(row=12, column=0, columnspan=5)
        self.quit_button.grid(row=13, column=0, rowspan=10, columnspan=10)
        
        # 3.2.3 Grades announcment
        if self.sound_variable.get() != "":
            GUI_Exam.speak(tell_grade(self.grade.get()))

        self.store_data()
        self.make_pdf()
        self.make_excel_summary()
    
    def for_correct_answer(self):
        """Provide a random message for correct answers."""
        return random.choice([
        "Bingo! You're practically a math magician!",
        "Nailed it! You're as sharp as a ninja star!",
        "Absolutely correct! You're a math superhero in the making!",
        "Epic win! You're the ruler of numbers!",
        "Bravo! You're a math wizard in the making!",
        "Fantastic! You're a superstar!",
        "Hooray! You did it!",
        "Congratulations! You're amazing!",
        "Great job! Keep up the good work!",
        "Wow! You're a champion!",
        "Superb effort! Well done!",
        "Bravo! You're a shining star!",
        "Congratulations on your success!",
        "Awesome work! You're a rockstar!",
        "Hip, hip, hooray! You're a winner!",
        "Well done! You make us proud!",
        "Congratulations! You're on fire!",
        "Incredible! Keep reaching for the stars!",
        "You did it! You're a real trooper!",
        "Cheers to your success! You're awesome!",
        "Woo-hoo! You're a success story!",
        "Way to go! You're a true hero!",
        "Congratulations, superstar! You're unstoppable!",
        "High fives! You're a fantastic friend!",
        "You did it with style! Congratulations!"
    ])
    
    def for_incorrect_answer(self):
        """Provide a random message for incorrect answers."""
        return random.choice([
        "Oopsie-doodle! No worries, superheroes stumble too!",
        "Close, but no cookie this time! You'll get it next round, I believe in you!",
        "Uh-oh! The numbers did a little dance, but don't worry, you'll catch the rhythm next time!",
        "Not quite, but you're on the right track! Keep up the awesome effort!",
        "Almost there! Your brain is flexing its muscles. Let's give it another shot!",
        "No worries! Mistakes help us learn. You've got this!",
        "Oops, that's okay! Keep trying, you'll get it right!",
        "Don't give up! You're getting closer with each attempt.",
        "Learning is an adventure. Keep exploring!",
        "It's okay to make mistakes. You're on the path to success!",
        "Every great scientist started with a few mistakes. You're a little scientist!",
        "Mistakes are proof that you are trying. Keep it up!",
        "You're doing fantastic! Keep going, you'll crack it!",
        "Great effort! You're making progress!",
        "Mistakes are just opportunities to learn something new. Well done!",
        "Keep that positive attitude! You're doing amazing things!",
        "You're a problem-solving champion in the making!",
        "Every mistake is a step closer to success. You're doing great!",
        "Learning is a journey, and you're on the right path!",
        "Believe in yourself! You're capable of incredible things.",
        "It's okay to struggle. That's how we become stronger!",
        "Perseverance is the key to success. You're persevering!",
        "Your efforts are commendable! Keep pushing forward.",
        "You're on a learning adventure! Keep up the good work.",
        "Remember, even the best had to practice. You're doing awesome!"
    ])

    def for_failed_attempt(self):
        """Provide a random message for failed attempts."""
        return random.choice([
        "Phew, tricky one! No worries, every mistake is a chance to learn something new!",
        "Whoa, that one did a little twist! Mistakes happen, but so does progress. Ready for the next adventure?",
        "That was a toughie! Don't worry, you're building a super-strong brain by giving it a workout!",
        "Not this time, but your determination is shining bright! Take a breath, and let's tackle the next challenge together!",
        "Whoopsie-daisy! Even the best explorers take a wrong turn. Shake it off, and let's set sail for the next discovery!",
        "That's okay! Don't worry, you gave it your best shot.",
        "No problem! Mistakes happen. You'll do better next time!",
        "Great effort! Remember, mistakes are stepping stones to success.",
        "You're resilient! Keep a positive attitude for the next challenge.",
        "Well done for trying! Learning is a journey, and you're on it!",
        "Fantastic attempt! Even the experts were once beginners.",
        "You're a superstar! Keep practicing, and you'll master it.",
        "It's okay to feel frustrated. Take a deep breath and try again later!",
        "You're a champion for giving it your all. Keep up the good work!",
        "Every mistake is a lesson learned. You're becoming wiser!",
        "You're on the right track! Keep going, and success will follow.",
        "Mistakes are proof that you're trying. Keep that positive spirit!",
        "You're making progress every day. Celebrate the small victories!",
        "The journey of learning is full of twists and turns. You're doing great!",
        "Perseverance is your superpower! Keep pushing forward.",
        "You've got the courage to face challenges. Keep that bravery!",
        "Your effort is what matters most. Keep up the hard work!",
        "Don't be discouraged! You're growing with every attempt.",
        "Remember, even the strongest heroes faced setbacks. You're a hero!",
        "Believe in yourself! Your potential is limitless."
    ])

    def store_data(self):
        if self.question_asked == 1 and self.test_end == None:
            self.file_open_mode = "w"
        else:
            self.file_open_mode = "a"
        if self.test_end == None:
            a = self.display_question.get()
            if self.question_paper._S == "fraction" and self.question_paper.choices:
                b = f"Your Answer: Option {self.choice_var.get()+1}"
            else:
                b = f"Your Answer: {self.input_user_answer.get()}"
            c = f"Remainder: {self.input_user_answer_remainder.get()}" if self.question_paper._S == "/" else None
            d = (
                f"You answered Correctly in Attempt No.: {self.attempts_counter+1}"
                if self.attempts_counter < 3
                else f"You answered this question incorrectly!"
            )
        else:
            a = f"Score: {self.exam_score}\nTotal Questions: {self.question_asked}\nPercent Marks: {round(self.exam_score/self.question_asked*100, 2)}%.\n{self.grade.get()}"
            b = f"Test Dated: {self.start_time.strftime('%d-%B-%Y')}\nTest Started: {self.test_start}"
            c = f"Test Ended: {self.test_end}"
            d = f"Exam Duration: {round((self.end_time - self.start_time).total_seconds()/60, 2)} minutes"
        with open(os.path.join(OUTPUT_DIR, f"{self.file_name}.txt"), (self.file_open_mode)) as file:
            if a!= None and b==None and c==None and d==None:
                file.write(str(f"{a}\n\n"))
                return
            elif a!= None and b!=None and c==None and d==None:
                file.write(str(f"{a}\n"))
                file.write(str(f"{b}\n\n"))
                return
            elif a!= None and b!=None and c!=None and d==None:
                file.write(str(f"{a}\n"))
                file.write(str(f"{b}\n"))
                file.write(str(f"{c}\n\n"))
                return
            elif a!= None and b!=None and c==None and d!=None:
                file.write(str(f"{a}\n"))
                file.write(str(f"{b}\n"))
                file.write(str(f"{d}\n\n"))
                return
            elif a!= None and b!=None and c!=None and d!=None:
                file.write(str(f"{a}\n"))
                file.write(str(f"{b}\n"))
                file.write(str(f"{c}\n"))
                file.write(str(f"{d}\n\n"))
                return

    def make_pdf(self):
        self.pdf = PDF()
        self.pdf.set_title("Mathematics Practice")
        self.pdf.set_author("Vijendra Singh")
        self.pdf.print_chapter(f"{self.file_name}.txt")
        self.pdf.output(os.path.join(OUTPUT_DIR, f"Worksheet_{datetime.now().strftime('%d-%b-%y-%I%M')}.pdf"))

    def update_difficulty_scores(self):
        """Adjust difficulty scores based on performance stats."""
        for op, data in self.stats.items():
            total = data.get("total_questions", 0)
            if total == 0:
                continue
            avg_time = data.get("total_time", 0) / total
            avg_attempts = data.get("total_attempts", 0) / total
            first_try_rate = data.get("first_try_correct", 0) / total
            delta = 0.05 * avg_time + 0.5 * (avg_attempts - 1) - 0.3 * first_try_rate
            difficulty_scores[op] = difficulty_scores.get(op, 2.0) + delta

    def make_excel_summary(self):
        self.update_difficulty_scores()

        rows = []
        for k, v in self.stats.items():
            total = v["total_questions"]
            correct = v["correct_answers"]
            attempts = v["total_attempts"]
            accuracy = round((correct / total * 100) if total else 0, 2)
            rows.append({
                "Question Type": op_names.get(k, k),
                "Total Questions": total,
                "Correct Answers": correct,
                "Total Attempts": attempts,
                "Accuracy (%)": accuracy,
            })

        df_summary = pd.DataFrame(rows)

        meta = {
            "Start Time": self.test_start,
            "End Time": self.test_end,
            "Duration": f"{round((self.end_time - self.start_time).total_seconds()/60, 2)} minutes",
            "Total Questions": self.question_asked,
            "Total Correct": self.exam_score,
            "Overall Accuracy (%)": round(self.exam_score / self.question_asked * 100, 2) if self.question_asked else 0,
        }
        meta_df = pd.DataFrame(list(meta.items()), columns=["Metric", "Value"])

        # update the master workbook with a new summary sheet and log entries
        self.update_all_sessions_log(df_summary, meta_df)

    def update_all_sessions_log(self, df_summary, meta_df):
        """Create or update AllSessions.xlsx with log, index, and summary sheets."""
        path = os.path.join(OUTPUT_DIR, "AllSessions.xlsx")
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
            # remove default sheet and create Log and Index
            default = wb.active
            wb.remove(default)
            log_ws = wb.create_sheet("Log")
            log_ws.append([
                "Date",
                "Time",
                "Question Type",
                "Total Questions",
                "Correct Answers",
                "Total Attempts",
                "Accuracy (%)",
                "Start Time",
                "End Time",
                "Duration",
            ])
            idx_ws = wb.create_sheet("Index")
            idx_ws.append([
                "Session Number",
                "Date",
                "Start Time",
                "End Time",
                "Duration",
                "Total Questions",
                "Accuracy (%)",
                "Summary Sheet",
            ])

        # ensure worksheets exist
        log_ws = wb["Log"] if "Log" in wb.sheetnames else wb.create_sheet("Log")
        idx_ws = wb["Index"] if "Index" in wb.sheetnames else wb.create_sheet("Index")

        from openpyxl.utils.dataframe import dataframe_to_rows

        # determine next session number based on summary sheets
        session_num = len([s for s in wb.sheetnames if s.startswith("Summary_")]) + 1
        summary_name = f"Summary_{session_num:03d}"
        summary_ws = wb.create_sheet(summary_name)

        # write summary table
        for r in dataframe_to_rows(df_summary, index=False, header=True):
            summary_ws.append(r)
        summary_ws.append([])
        for r in dataframe_to_rows(meta_df, index=False, header=True):
            summary_ws.append(r)

        # append entries to log sheet
        date_str = self.start_time.strftime("%Y-%m-%d")
        time_str = self.start_time.strftime("%H:%M:%S")
        start_full = self.start_time.strftime("%Y-%m-%d %H:%M:%S")
        end_full = self.end_time.strftime("%Y-%m-%d %H:%M:%S")
        duration = round((self.end_time - self.start_time).total_seconds() / 60, 2)

        for r in df_summary.to_dict("records"):
            log_ws.append([
                date_str,
                time_str,
                r["Question Type"],
                r["Total Questions"],
                r["Correct Answers"],
                r["Total Attempts"],
                r["Accuracy (%)"],
                start_full,
                end_full,
                duration,
            ])

        # append index row with hyperlink to summary sheet
        overall_accuracy = meta_df.loc[meta_df["Metric"] == "Overall Accuracy (%)", "Value"].iloc[0]
        idx_row = [
            session_num,
            date_str,
            time_str,
            self.end_time.strftime("%H:%M:%S"),
            duration,
            self.question_asked,
            overall_accuracy,
        ]
        idx_ws.append(idx_row + [summary_name])
        link_cell = idx_ws.cell(row=idx_ws.max_row, column=len(idx_row) + 1)
        link_cell.hyperlink = f"#{summary_name}!A1"
        link_cell.style = "Hyperlink"

        # save workbook and then record difficulty scores for this session
        wb.save(path)
        append_difficulty_session(difficulty_scores)

        save_difficulty_scores(difficulty_scores)


class PDF(FPDF, GUI_Exam):
    def header(self):
        logo_path = resource_path("logo_image.jpg")
        if os.path.exists(logo_path):
            # Rendering logo if available
            self.image(logo_path, 10, 8, 15)
        # Setting font: helvetica bold 15
        self.set_font("helvetica", "B", 15)
        # Calculating width of title and setting cursor position:
        width = self.get_string_width(self.title) + 6
        self.set_x((210 - width) / 2)
        # Setting colors for frame, background and text:
        self.set_draw_color(0, 80, 180)
        self.set_fill_color(230, 230, 0)
        self.set_text_color(220, 50, 50)
        # Setting thickness of the frame (1 mm)
        self.set_line_width(1)
        # Printing title:
        self.cell(
            width,
            9,
            self.title,
            border=1,
            align="C",
            fill=True,
        )
        # Performing a line break:
        self.ln(15)
    
    def footer(self):
        # Setting position at 1.5 cm from bottom:
        self.set_y(-15)
        # Setting font: helvetica italic 8
        self.set_font("helvetica", "I", 8)
        # Setting text color to gray:
        self.set_text_color(128)
        # Printing page number
        self.cell(0, 10, f"Page {self.page_no()}", align="C")

    def chapter_body(self, filepath):
        # Reading text file:
        with open(os.path.join(OUTPUT_DIR, f"{root_instance.file_name}.txt"), "rb") as fh:
            txt = fh.read().decode("latin-1")
        # Setting font: Times 12
        self.set_font("Times", size=12)
        # Printing justified text:
        self.multi_cell(0, 5, txt)
        # Performing a line break:
        self.ln()
        # Final mention in italics:
        self.set_font("helvetica", "I", 8)
        self.cell(0, 5, "(End of test!)", align="C")

    def print_chapter(self, filepath):
        self.add_page()
        self.chapter_body(filepath)


def main():
    global root_instance
    root_instance = GUI_Exam.launch_main()
    GUI_Exam.root.mainloop()

def get_grade(m, t):
    """
    Calculate the user's grade based on the exam score.

    Parameters:
        - m: User's exam score
        - t: Total number of questions

    Returns:
        - A string representing the user's grade
    """
    score = (m/t)*100
    if score >= 90:
        return("Grade: A")
    elif score >= 80:
        return("Grade: B")
    elif score >= 70:
        return("Grade: C")
    elif score >=60:
        return("Grade: D")
    else:
        return("Grade: F")

def evaluate(answer_user, answer_actual, answer_user_remainder=None, answer_actual_remainder=None, sign=None):
    if sign != "/":
        if answer_user == answer_actual:
            return True
        else:
            return False
    else:
        if answer_user == answer_actual and answer_user_remainder == answer_actual_remainder:
            return True
        else:
            return False

def factors_of(n: int):
    return [i for i in range(1, n + 1) if n % i == 0]

def is_prime(n: int) -> bool:
    if n < 2:
        return False
    for i in range(2, int(n ** 0.5) + 1):
        if n % i == 0:
            return False
    return True

def twin_prime_pair(n: int):
    if not is_prime(n):
        return None
    if n - 2 >= 2 and is_prime(n - 2):
        return (n - 2, n)
    if n + 2 <= 100 and is_prime(n + 2):
        return (n, n + 2)
    return None


def prime_factorization(n: int):
    """Return the list of prime factors for *n* including multiplicities."""
    factors = []
    divisor = 2
    while divisor * divisor <= n:
        while n % divisor == 0:
            factors.append(divisor)
            n //= divisor
        divisor += 1
    if n > 1:
        factors.append(n)
    return factors

def lcm_of_numbers(numbers):
    from math import gcd
    lcm_val = numbers[0]
    for n in numbers[1:]:
        lcm_val = lcm_val * n // gcd(lcm_val, n)
    return lcm_val

def lcm_explanation(nums, method):
    lcm_val = lcm_of_numbers(nums)
    if method == "listing multiples":
        parts = []
        for n in nums:
            multiples = [n * i for i in range(1, lcm_val // n + 1)]
            parts.append(f"Multiples of {n}: {', '.join(map(str, multiples))}")
        return f"{' ; '.join(parts)}. The first common multiple is {lcm_val}."
    elif method == "prime factorization":
        pf_texts = [f"Prime factors of {n}: {' × '.join(map(str, prime_factorization(n)))}" for n in nums]
        return f"{' ; '.join(pf_texts)}. Multiply highest powers of each prime to get {lcm_val}."
    else:
        temps = nums[:]
        prime = 2
        steps = []
        factors = []
        while any(t > 1 for t in temps):
            divided = False
            for i in range(len(temps)):
                if temps[i] % prime == 0:
                    temps[i] //= prime
                    divided = True
            if divided:
                factors.append(str(prime))
                steps.append(f"divide by {prime} → {', '.join(map(str, temps))}")
            else:
                prime += 1
        return f"{' ; '.join(steps)}. Multiply {', '.join(factors)} to get {lcm_val}."


def parse_factor_input(text: str):
    """Parse user entered prime factors separated by ×, *, or spaces."""
    parts = [p for p in re.split(r"[×*\s]+", text.strip()) if p]
    if not parts:
        return None
    if not all(part.isdigit() for part in parts):
        return None
    return [int(p) for p in parts]


def parse_fraction_input(text: str):
    """Parse a fraction like 'a/b' or a whole number."""
    text = text.strip()
    if text.isdigit():
        return int(text), 1
    m = re.match(r"^(\d+)\s*/\s*(\d+)$", text)
    if not m:
        return None
    num, den = int(m.group(1)), int(m.group(2))
    if den == 0:
        return None
    return num, den


def tell_grade(grade):
    """Provide a random congratulatory message based on the grade."""
    if grade == "Grade: A":
        return random.choice([
    "Fantastic! You've reached Grade A! You're a math wizard in the making!",
    "Incredible job! Grade A is the highest honor, and you've earned it with your exceptional skills.",
    "Wow! You're a mathematical genius! Grade A is a testament to your brilliance.",
    "Amazing! Your Grade A achievement proves that you're a math superhero!",
    "Brilliant work! Grade A means you've mastered the math quest. Keep up the fantastic effort!"
])
    elif grade == "Grade: B":
        return random.choice([
    "Bravo! Grade B is outstanding! Keep up the great work, you're mastering these math challenges!",
    "Impressive! Grade B showcases your dedication to excellence in math.",
    "Well done! Grade B is a mark of your commitment and hard work in the math adventure.",
    "Great job! You've achieved Grade B, and you're on the path to becoming a math star!",
    "Excellent effort! Grade B reflects your strong performance in the math quest. Keep shining!"
])
    elif grade == "Grade: C":
        return random.choice([
    "Congratulations on achieving Grade C! You're doing well, and with a bit more practice, you'll shine even brighter!",
    "Good work! Grade C signifies your steady progress in mastering math skills.",
    "Well deserved! Grade C shows your commitment to learning and improvement.",
    "Thumbs up! Grade C is a positive step, and you're on your way to conquering more math challenges.",
    "Keep it up! Grade C is a commendable achievement, and you're on the right track!"
])
    elif grade == "Grade: D":
        return random.choice([
    "Great effort! Grade D shows progress, and you're on the right track. Keep practicing, and you'll see amazing results!",
    "Well done on achieving Grade D! Your dedication is paying off, and you're improving in math.",
    "Persistence pays off! Grade D acknowledges your hard work and commitment to overcoming math obstacles.",
    "You're getting there! Grade D is a step forward, and you're making strides in the math adventure.",
    "Good job! Grade D recognizes your efforts, and you're making progress in the world of math."
])
    elif grade == "Grade: F":
        return random.choice([
    "No worries! Even superheroes face challenges. Grade F is just a stepping stone. With persistence, you'll conquer every math quest!",
    "Keep going! Grade F is a chance to learn and grow. You'll overcome math challenges with determination.",
    "Every setback is a setup for a comeback! Grade F is a starting point, and you'll rise to new math heights.",
    "Stay positive! Grade F is an opportunity to improve and become an even better math explorer.",
    "You're on a math journey, and Grade F is a part of the adventure. Keep going, and you'll achieve great things!"
])


if __name__ == "__main__":
    main()
